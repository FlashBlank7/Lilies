"""E2 evidence for SCENARIO P1-1: raw tabular data -> clean Excel artifact.

Real CSV -> platform parse_table (real CSV parsing) -> workflow
(start -> record_collection_normalize -> record_deduplicate ->
 variable_assigner(builds spec) -> typed_workbook -> end)
-> run -> real .xlsx artifact.
"""
import json
import os
import sys
import time
import zipfile
from pathlib import Path

sys.path.insert(0, "/home/jiangzhijun/Lilies/platform/backend/src")

from fastapi.testclient import TestClient

from agent_platform.api import create_app
from agent_platform.config import Settings
from agent_platform.table_intake import parse_table

CSV_PATH = "/tmp/lilies_p1_sales.csv"
DATA_DIR = "/tmp/lilies_p1_data"
WORKSPACE_ROOT = Path("/tmp/lilies_p1_workspaces")

TOKEN = "workflow-test"
H = {"Authorization": f"Bearer {TOKEN}"}

# ---------------------------------------------------------------- 1. CSV + parse
raw_bytes = Path(CSV_PATH).read_bytes()
intake = parse_table(CSV_PATH, data=raw_bytes)
rows = intake["rows"]
print(f"[intake] columns={intake['columns']}")
print(f"[intake] raw rows parsed from real CSV = {len(rows)}")
unique_ids = len({r["order_id"] for r in rows})
print(f"[intake] distinct order_id = {unique_ids}")
print(f"[intake] sample row0 = {rows[0]}")

# ---------------------------------------------------------------- 2. App + workflow
settings = Settings(
    api_token=TOKEN,
    model_egress_enabled=True,
    data_dir=DATA_DIR,
    workspace_root=str(WORKSPACE_ROOT),
)
app = create_app(settings)  # NO provider -> no model egress; workflow is LLM-free

with TestClient(app) as client:
    # create application
    r = client.post(
        "/api/v1/applications",
        headers=H,
        json={
            "name": "P1-1 Sales Data Pipeline",
            "description": "CSV -> clean -> Excel artifact E2 evidence",
            "requirement": "Turn raw tabular sales data into a clean typed Excel artifact.",
        },
    )
    print(f"[app] create status={r.status_code}")
    app_id = r.json()["id"]
    print(f"[app] application_id={app_id}")

    draft = client.get(f"/api/v1/applications/{app_id}/draft", headers=H).json()
    rev = int(draft["revision"])
    print(f"[draft] initial revision={rev}")

    nodes = [
        {
            "id": "start",
            "type": "start",
            "title": "Sales Input",
            "config": {
                "inputs": [
                    {"name": "records", "label": "Sales records", "type": "array", "required": True}
                ]
            },
        },
        {
            "id": "normalize",
            "type": "record_collection_normalize",
            "title": "Normalize Envelope",
            "config": {
                "value": {"$ref": {"node_id": "start", "path": ["records"]}},
            },
        },
        {
            "id": "dedupe",
            "type": "record_deduplicate",
            "title": "Deduplicate Orders",
            "config": {
                "records": {"$ref": {"node_id": "normalize", "path": ["records"]}},
                "key_paths": [["order_id"]],
                "missing_key_policy": "error",
            },
        },
        {
            "id": "assigner",
            "type": "variable_assigner",
            "title": "Build Workbook Spec",
            "config": {
                "assignments": {
                    "workbook": {
                        "sheets": [
                            {
                                "name": "Clean Sales",
                                "columns": [
                                    {"key": "order_id", "header": "Order ID", "type": "string"},
                                    {"key": "customer", "header": "Customer", "type": "string"},
                                    {"key": "product", "header": "Product", "type": "string"},
                                    {"key": "category", "header": "Category", "type": "string"},
                                    {"key": "units", "header": "Units", "type": "integer"},
                                    {"key": "unit_price", "header": "Unit Price", "type": "number"},
                                    {"key": "revenue", "header": "Revenue", "type": "number"},
                                    {"key": "order_date", "header": "Order Date", "type": "date"},
                                ],
                                "rows": {"$ref": {"node_id": "dedupe", "path": ["unique"]}},
                            }
                        ]
                    }
                }
            },
        },
        {
            "id": "workbook",
            "type": "typed_workbook",
            "title": "Write Excel Artifact",
            "config": {
                "spec": {"$ref": {"node_id": "assigner", "path": ["output", "workbook"]}},
                "filename": "clean_sales.xlsx",
                "formula_policy": "reject",
                "lineage": [
                    {"source_type": "workflow_input", "reference": "records"},
                    {"source_type": "node_output", "reference": "dedupe.unique"},
                ],
            },
        },
        {
            "id": "end",
            "type": "end",
            "title": "Done",
            "config": {
                "outputs": {
                    "workbook_artifact": {"$ref": {"node_id": "workbook", "path": ["artifact"]}},
                    "dedup_receipt_count": {"$ref": {"node_id": "dedupe", "path": ["receipts", 0]}},
                }
            },
        },
    ]

    edges = [
        {"source": "start", "target": "normalize"},
        {"source": "normalize", "target": "dedupe", "source_port": "records", "target_port": "input"},
        {"source": "dedupe", "target": "assigner"},
        {"source": "assigner", "target": "workbook"},
        {"source": "workbook", "target": "end"},
    ]

    for i, node in enumerate(nodes):
        op = {
            "expected_revision": rev,
            "idempotency_key": f"p1-1-add-node-{node['id']}-{i}",
            "op": "add_node",
            "data": {"node": node},
        }
        r = client.post(f"/api/v1/applications/{app_id}/draft", headers=H, json=op)
        if r.status_code != 200:
            print(f"[draft] add_node {node['id']} FAILED: {r.status_code} {r.text}")
            sys.exit(2)
        rev = int(r.json()["revision"])
        print(f"[draft] added node {node['id']} -> revision {rev}")

    for j, edge in enumerate(edges):
        op = {
            "expected_revision": rev,
            "idempotency_key": f"p1-1-add-edge-{j}",
            "op": "add_edge",
            "data": {"edge": edge},
        }
        r = client.post(f"/api/v1/applications/{app_id}/draft", headers=H, json=op)
        if r.status_code != 200:
            print(f"[draft] add_edge {edge.get('source')}->{edge.get('target')} FAILED: {r.status_code} {r.text}")
            sys.exit(2)
        rev = int(r.json()["revision"])
        print(f"[draft] added edge {edge.get('source')}->{edge.get('target')} -> revision {rev}")

    # validate draft
    val = client.post(
        f"/api/v1/applications/{app_id}/draft/validate", headers=H, json={}
    ).json()
    print(f"[draft] validate valid={val.get('valid')} errors={val.get('errors')}")

    final = client.get(f"/api/v1/applications/{app_id}/draft", headers=H).json()
    wf = final["snapshot"]["workflow"]
    print(f"[draft] final nodes={[n['id'] for n in wf['nodes']]} edges={[(e['source'], e['target']) for e in wf['edges']]}")

    # ---------------------------------------------------------------- 3. Run
    run = client.post(
        f"/api/v1/applications/{app_id}/runs",
        headers=H,
        json={"inputs": {"records": rows}, "use_draft": True},
    )
    print(f"[run] create status={run.status_code}")
    if run.status_code != 202:
        print(f"[run] create failed: {run.text}")
        sys.exit(3)
    run_id = run.json()["run_id"]
    print(f"[run] run_id={run_id}")

    status = None
    for _ in range(60):
        rr = client.get(f"/api/v1/runs/{run_id}", headers=H).json()
        status = rr.get("status")
        if status in ("succeeded", "failed", "cancelled", "paused"):
            break
        time.sleep(0.3)
    print(f"[run] final status={status}")
    if status != "succeeded":
        print(f"[run] state: {json.dumps(rr, default=str)[:2000]}")
        sys.exit(4)

    outputs = rr.get("outputs") or {}
    artifact = outputs.get("workbook_artifact") or {}
    print(f"[run] outputs keys={list(outputs.keys())}")
    print(f"[run] artifact = {json.dumps(artifact, default=str)[:600]}")

    # ---------------------------------------------------------------- 4. Verify artifact on disk + via API
    on_disk = WORKSPACE_ROOT / ".workflow-run-artifacts" / run_id / "artifacts" / "clean_sales.xlsx"
    print(f"[artifact] on-disk exists={on_disk.exists()} size={on_disk.stat().st_size if on_disk.exists() else 'N/A'}")
    if not on_disk.exists():
        print("[artifact] NOT FOUND on disk; listing run artifact dir")
        for p in (WORKSPACE_ROOT / ".workflow-run-artifacts" / run_id).rglob("*"):
            print("   ", p.relative_to(WORKSPACE_ROOT))
        sys.exit(5)

    # API artifacts endpoint
    code = client.get(f"/api/v1/applications/{app_id}/access-code", headers=H).json()["code"]
    arts = client.get(f"/api/v1/use/{app_id}/runs/{run_id}/artifacts", params={"code": code}).json()
    print(f"[artifact] /use artifacts endpoint -> {arts}")

    # structural xlsx validation
    is_zip = zipfile.is_zipfile(on_disk)
    names = zipfile.ZipFile(on_disk).namelist() if is_zip else []
    print(f"[artifact] is_valid_zip={is_zip} members={len(names)}")
    print(f"[artifact] has workbook part={'xl/workbook.xml' in names} sheet_count={len([n for n in names if n.startswith('xl/worksheets/')])}")

    # openpyxl round-trip
    import openpyxl
    wb = openpyxl.load_workbook(on_disk, data_only=True)
    ws = wb["Clean Sales"]
    header = [c.value for c in ws[1]]
    data_rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2, values_only=False)]
    data_rows = [row for row in data_rows if any(v is not None for v in row)]
    print(f"[artifact] openpyxl sheets={wb.sheetnames}")
    print(f"[artifact] header={header}")
    print(f"[artifact] data row count in xlsx = {len(data_rows)}")
    print(f"[artifact] first data row = {data_rows[0] if data_rows else None}")
    order_ids_in_xlsx = [row[0] for row in data_rows]
    print(f"[artifact] xlsx order_ids unique = {len(set(order_ids_in_xlsx))}")

    assert len(data_rows) == unique_ids, "xlsx row count != unique order_ids"
    assert len(data_rows) == 10
    print("[verify] PASS: xlsx contains exactly the deduplicated unique records")

    # node execution order + artifact.created events (JSON list via /v1/streams)
    ev = client.get(f"/v1/streams/{run_id}", headers=H).json()
    print("[events] node execution order:")
    for e in ev:
        k = e.get("kind")
        if k == "node.started":
            d = e.get("data", {})
            print(f"    node.started node_id={d.get('node_id')} type={d.get('type')}")
        elif k == "artifact.created":
            print(f"    artifact.created media_type={e['data'].get('media_type')} path={e['data'].get('relative_path')} sha256={e['data'].get('sha256')}")
    print(f"[run] dedup receipt 0 (first receipt) = {json.dumps(outputs.get('dedup_receipt_count'), default=str)}")
